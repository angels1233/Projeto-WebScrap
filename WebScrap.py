import pandas as pd
from datetime import datetime
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

#Função que realiza a busca dentro do site
def Pesquisa(url, pesquisa):

   #variáveis que será usada pra criar o dicionário 
    preco_old_list = []
    preco_new_list = []
    desconto_list = []
    link_list = []
    nome_list = []

    #Iniciando o webdriver do google e atualizando o drive do mesmo.
    service = Service(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    options.add_argument("--log-level=3")
    driver = webdriver.Chrome(service = service, options = options)

    #abrindo o navegador no site de busca selecionado
    driver.get(url)

    #Colocando a elemento da pesquisa no pesquisador do site
    driver.find_element(By.ID,'cb1-edit').send_keys(pesquisa)
    time.sleep(1)

    #Acionando o motor de busca do site
    driver.find_element(By.CLASS_NAME,'nav-search-btn').click()
    time.sleep(3)
    
    #Criando uma lista de itens padronizada de acordo com o site de busca
    itens = driver.find_elements(By.CLASS_NAME,'ui-search-layout__item')

    #For que percorrer os itens pra pegar os elementos de cada um deles
    for iten in itens:
        
        #Pega o url do item
        link = iten.find_element(By.TAG_NAME,'a').get_attribute('href')
        #Pega o nome do item 
        nome_aux = iten.find_element(By.CLASS_NAME,'poly-component__title-wrapper')
        nome = nome_aux.find_element(By.CLASS_NAME,'poly-component__title').text
        #Pega o preço do item
        preco_new_aux = iten.find_element(By.CLASS_NAME,'poly-price__current')    
        preco_new = preco_new_aux.find_element(By.CLASS_NAME,'andes-money-amount__fraction').text
        
        #Pega o valro dos centavos se ouver e caso não adiciona o valor zerado
        try:
            cents_new = preco_new_aux.find_element(By.CLASS_NAME,'andes-money-amount__cents').text
            preco_new = (preco_new +',' + cents_new)
        
        except:
             preco_new = (preco_new + ',' + '00')

        #Caso o produto possua desconto pega o valor em procentagem do desconto e o valor sem o desconto
        try:
            iten.find_element(By.CLASS_NAME,'andes-money-amount--previous')
            desconto = preco_new_aux.find_element(By.CLASS_NAME,'andes-money-amount__discount').text[:3]
            preco_old_aux = iten.find_element(By.CLASS_NAME,'andes-money-amount--previous')
            preco_old = preco_old_aux.find_element(By.CLASS_NAME,'andes-money-amount__fraction').text
            
            try:
                cents_old = preco_old_aux.find_element(By.CLASS_NAME,'andes-money-amount__cents').text
                preco_old = (preco_old + ',' + cents_old)
            except:
                preco_old = (preco_old + ',' + '00')

        except:
            preco_old = ''
            desconto = ''

        #Coloca as informações no numa lista
        preco_old_list.append(preco_old)
        preco_new_list.append(preco_new)
        desconto_list.append(desconto)
        link_list.append(link)
        nome_list.append(nome)

        print('preço full: '+ preco_old + '  Desconto: ' + desconto + '  preço com desconto: ' + preco_new)

    #Formata as informações no dicionário    
    dictIten = {'nome': nome_list,
                'preco': preco_new_list,
                'desconto': desconto_list,
                'preco_anterior': preco_old_list,
                'link': link_list}
    
    # return pd.DataFrame.from_dict(dictIten)
    return dictIten
    time.sleep(10)

#função que tranforma o dicionário em uma tabela do excel
def Tabela_pesquisa(procura):
   
   #Onde está definido o site de busca e retorna os intens da pesquisa
   df_pesquisa = Pesquisa('https://www.mercadolivre.com.br/', procura) 

   #Cria o arquivo do excel vazio com uma aba. 
   arquivo = Workbook()
   aba = arquivo.active
   aba.title = "Base de dados"

   #Retorna os nomes do cabeçalho do dicionário
   nomes_colunas = list(df_pesquisa.keys())

   #Iteração com os intens do dicionário na tabela do excel
   for num_coluna, nome_coluna in enumerate(nomes_colunas, start=1):
       celula = aba.cell(row=1, column=num_coluna)
       celula.value = nome_coluna
       celula.alignment = Alignment(horizontal='center')
       lista_valores = df_pesquisa[nome_coluna]
       for num_linha, valor, in enumerate(lista_valores, start=2):
           celula = aba.cell(row=num_linha, column=num_coluna)
           celula.value = valor
           celula.alignment = Alignment(horizontal='center')
           if num_coluna == 5:
            celula.hyperlink = valor
            celula.alignment = Alignment(horizontal="left")

   #Salva a tabela do excel na pasta correta com o nome do item pesquisado e data da pesquisa
   caminho_pasta = 'Projeto-WebScrap\\Pesquisados\\'
   data = (str(datetime.now()))
   nome = (procura + '_' + data[:10] +'.xlsx')
   arquivo.save(caminho_pasta + nome)



print('O que deseja pesquisar: ')
pesquisado = input()

Tabela_pesquisa(pesquisado)